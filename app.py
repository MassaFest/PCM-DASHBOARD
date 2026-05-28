import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta
import requests
import re
import unicodedata
from io import StringIO, BytesIO

st.set_page_config(
    page_title="PCM MassaFest",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── Paleta industrial laranja ──────────────────────────────────
COR_PRIMARIA   = "#F97316"   # laranja
COR_PERIGO     = "#ef4444"   # vermelho
COR_AVISO      = "#eab308"   # amarelo
COR_OK         = "#22c55e"   # verde
COR_INFO       = "#3b82f6"   # azul
ESCALA_LARANJA = ["#1a0a00","#7c2d00","#c2410c","#f97316","#fdba74","#fff7ed"]
ESCALA_CALOR   = "YlOrRd"

DEFAULT_SHEET_CHAMADOS  = "1AJXsR6YpgSuYrDRo_vtNpsJtn0vSQ9ayjBX_4mMTtdw"
DEFAULT_GID_CHAMADOS    = "2143366097"
DEFAULT_SHEET_RETORNOS  = "1KO_U-Ly1s9rW2YuPdc48zwOcR6EuPf1CfBGbbSZMjEY"
DEFAULT_GID_RETORNOS    = "824431047"
DEFAULT_SHEET_CAT1      = "1gTmX6wTU2KBuI-dg2HHDTH-xTSNSkZXOUdQVpFmp-oo"
DEFAULT_SHEET_CAT2      = "1bxr1yw-DcYrExfRpUYMEjf4CAm7uWe9zbDcTxB3wMfs"
DEFAULT_SHEET_PREVENTIVA = ""   # usuário irá configurar

@st.cache_data(ttl=3600)
def load_catalogo(sheet_id: str) -> tuple[dict, pd.DataFrame]:
    """
    Lê todas as abas de uma planilha de catálogo de máquinas.
    Retorna:
      - catalogo: dict {codigo_norm → {nome, localizacao}}
      - historico: DataFrame com histórico legado de todas as máquinas
    Estrutura esperada em cada aba:
      Linha com 'Patrimônio' | 'Nome' | 'Localização'  (cabeçalho)
      Próxima linha          | dados da máquina
      ...
      Linha com 'Data' | 'Tipo de manutenção' | ...    (cabeçalho histórico)
      linhas de histórico...
    """
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    catalogo  = {}
    hist_rows = []

    try:
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
        xls = pd.read_excel(BytesIO(resp.content), sheet_name=None, header=None)

        for tab_name, df in xls.items():
            df = df.fillna("").astype(str)
            codigo_maq = nome_maq = local_maq = ""

            # ── Procura linha de cabeçalho do cadastro ──────────────────────
            for i, row in df.iterrows():
                vals = [v.strip().lower() for v in row.values]
                if any("patrim" in v for v in vals):
                    if i + 1 < len(df):
                        dr = df.iloc[i + 1]
                        codigo_raw = dr.iloc[0].strip()
                        nome_maq   = dr.iloc[1].strip() if len(dr) > 1 else ""
                        local_maq  = dr.iloc[2].strip() if len(dr) > 2 else ""
                        # Normaliza "MP 003" → "MP003"
                        codigo_maq = re.sub(r'\s+', '', codigo_raw).upper()
                    break

            if not codigo_maq or not nome_maq or nome_maq.lower() in ("", "nan"):
                continue

            catalogo[codigo_maq] = {"nome": nome_maq, "localizacao": local_maq}

            # ── Procura histórico de manutenção na mesma aba ────────────────
            hist_start = None
            for i, row in df.iterrows():
                vals = [v.strip().lower() for v in row.values]
                if any("data" in v for v in vals) and any("manuten" in v for v in vals):
                    hist_start = i
                    break

            if hist_start is not None:
                hist_df = df.iloc[hist_start:].copy()
                hist_df.columns = range(len(hist_df.columns))
                # Primeira linha vira cabeçalho
                new_cols = [str(hist_df.iloc[0][c]).strip() for c in range(len(hist_df.columns))]
                hist_df  = hist_df.iloc[1:].copy()
                hist_df.columns = new_cols
                hist_df = hist_df[hist_df.iloc[:, 0].str.strip().astype(bool)]  # remove vazios
                hist_df = hist_df[~hist_df.iloc[:, 0].str.lower().str.contains("data")]

                # Renomeia colunas para padrão
                col_map = {}
                for c in hist_df.columns:
                    cl = c.lower()
                    if "data"    in cl: col_map[c] = "Data/Hora"
                    elif "tipo"  in cl: col_map[c] = "Tipo"
                    elif "troca" in cl or "pec" in cl: col_map[c] = "Troca de Peça"
                    elif "serv"  in cl or "realiz" in cl: col_map[c] = "Descrição"
                    elif "custo" in cl or "r$" in cl: col_map[c] = "Custo"
                hist_df = hist_df.rename(columns=col_map)
                hist_df["Máquina"]  = codigo_maq
                hist_df["Nome"]     = nome_maq
                hist_df["Origem"]   = "Histórico Planilha"
                if "Data/Hora" in hist_df.columns:
                    hist_df["Data/Hora"] = pd.to_datetime(
                        hist_df["Data/Hora"], dayfirst=True, errors="coerce"
                    )
                    hist_df = hist_df[hist_df["Data/Hora"].notna()]
                hist_rows.append(hist_df)

    except Exception as e:
        st.warning(f"Catálogo ({sheet_id[:20]}…): {e}")

    hist_legado = pd.concat(hist_rows, ignore_index=True) if hist_rows else pd.DataFrame()
    return catalogo, hist_legado


@st.cache_data(ttl=300)
def load_sheet(sheet_id: str, gid: str) -> pd.DataFrame:
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    try:
        resp = requests.get(url, timeout=15)
        resp.raise_for_status()
        df = pd.read_csv(StringIO(resp.text))
        df.columns = df.columns.str.strip()
        return df
    except Exception as e:
        st.error(f"Erro ao carregar planilha (id={sheet_id}, gid={gid}): {e}")
        return pd.DataFrame()

def parse_date_col(df: pd.DataFrame, col: str) -> pd.DataFrame:
    if col in df.columns:
        df[col] = pd.to_datetime(df[col], dayfirst=True, errors="coerce")
    return df

def limpar_nome_mecanico(valor: str) -> str:
    import re
    if not isinstance(valor, str):
        return valor
    nome = re.sub(r'^\d+\s*', '', valor.strip())   # números no início
    nome = re.sub(r'\s*\d+$', '', nome)              # números no fim
    nome = re.sub(r'\d+', '', nome)                  # números no meio
    nome = nome.strip().title()
    return nome if nome else ""

def normalizar_maquina(valor) -> str:
    """
    Aceita APENAS três formatos. Todo o resto é descartado.
      1. MPxxx / PAxxx  → ex: MP001, PA045  (novo padrão)
      2. Número puro     → ex: 001, 038, 13  (legado)
      3. Sem número      → qualquer variação → 'S/N'
    """
    import re, unicodedata

    if not isinstance(valor, str):
        return ""

    # Normaliza encoding (remove acentos para comparação)
    v = valor.strip()
    v_norm = unicodedata.normalize("NFD", v.lower())
    v_norm = re.sub(r'[̀-ͯ]', '', v_norm)  # remove diacríticos

    # ── Sem número → tudo vira S/N ──────────────────────────────────────────
    sem_numero = [
        r'^s\s*[/\\.]?\s*n\.?$',           # S/N, SN, S.N., s n
        r'^sem\s*nu?me?ro$',                # sem numero, sem nmero
        r'^na?o\s*tem$',                    # não tem, nao tem
        r'^sem\s*pat',                      # sem patrimônio
        r'^nao\s*ha$',                      # não há
        r'^nenhum[ao]?$',                   # nenhum
    ]
    for pat in sem_numero:
        if re.match(pat, v_norm):
            return "S/N"

    # ── Novo padrão: MP ou PA + 1 a 4 dígitos ───────────────────────────────
    if re.fullmatch(r'(?i)(MP|PA)\d{1,4}', v):
        return v.upper()

    # ── Legado: APENAS dígitos, 1 a 4 caracteres ────────────────────────────
    if re.fullmatch(r'\d{1,4}', v):
        num_norm = v.zfill(3)
        # Tenta converter para MP/PA via catálogo
        try:
            if num_norm in legacy_map:
                return legacy_map[num_norm]
            if v.lstrip('0') in legacy_map:
                return legacy_map[v.lstrip('0')]
        except NameError:
            pass  # legacy_map ainda não carregado
        return num_norm

    # ── Tudo o mais é inválido ───────────────────────────────────────────────
    return ""


def expandir_mecanicos(df: pd.DataFrame, col: str) -> pd.DataFrame:
    """
    Linhas com múltiplos mecânicos ("Lucas/Janailson", "Alexandre E Lucas")
    são expandidas: cada mecânico recebe uma linha individual.
    Linhas cujo valor seja apenas número (ex: "1950") são descartadas.
    """
    import re
    if col not in df.columns:
        return df

    linhas = []
    for _, row in df.iterrows():
        val = row[col]
        if not isinstance(val, str) or not val.strip():
            linhas.append(row)
            continue

        # Separadores: / | e | E | & | , | + (com ou sem espaços)
        partes = re.split(r'\s*/\s*|\s+[eE&]\s+|\s*,\s*|\s*\+\s*|\s*\|\s*', val)
        partes = [limpar_nome_mecanico(p) for p in partes]
        # Remove vazios e entradas que são só números
        partes = [p for p in partes if p and not re.fullmatch(r'\d+', p.strip())]

        if not partes:
            continue  # descarta linha sem nome válido

        for nome in partes:
            nova = row.copy()
            nova[col] = nome
            linhas.append(nova)

    return pd.DataFrame(linhas, columns=df.columns).reset_index(drop=True)

def agrupar_nomes_similares(series: pd.Series, threshold: float = 0.75) -> pd.Series:
    """
    Agrupa nomes parecidos (erros de digitação, variações) num único nome canônico.
    O nome canônico é o mais frequente do grupo.
    Ex: ['Alexandre','Alexadre','Alexander'] → todos viram 'Alexandre'
    """
    from difflib import SequenceMatcher

    contagem = series.value_counts()
    nomes    = list(contagem.index)
    mapa     = {}           # nome original → nome canônico
    visitado = set()

    for i, n1 in enumerate(nomes):
        if n1 in visitado:
            continue
        grupo = [n1]
        for n2 in nomes[i+1:]:
            if n2 in visitado:
                continue
            ratio = SequenceMatcher(None, n1.lower(), n2.lower()).ratio()
            if ratio >= threshold:
                grupo.append(n2)
                visitado.add(n2)
        # Canônico = o mais frequente do grupo
        canonico = max(grupo, key=lambda n: contagem.get(n, 0))
        for n in grupo:
            mapa[n] = canonico
        visitado.add(n1)

    return series.map(lambda x: mapa.get(x, x) if isinstance(x, str) else x)


def formatar_tempo(minutos):
    """Formata minutos em texto legível: ex. 1h 30min"""
    if pd.isna(minutos) or minutos < 0:
        return "—"
    h = int(minutos // 60)
    m = int(minutos % 60)
    if h > 0 and m > 0:
        return f"{h}h {m}min"
    elif h > 0:
        return f"{h}h"
    else:
        return f"{m}min"

def calcular_tempos(df_ch, df_ret, col_data_cham, col_maquina, col_data_ret, maq_col_ret):
    """
    Para cada retorno, busca o chamado mais recente da mesma máquina
    ANTES da data do retorno e calcula o tempo de resposta em minutos.
    Retorna df_ret com coluna 'Tempo_min' adicionada.
    """
    if (df_ch.empty or df_ret.empty
            or col_data_cham == "(nenhuma)" or col_maquina == "(nenhuma)"
            or col_data_ret == "(nenhuma)" or not maq_col_ret):
        return df_ret

    df_ret = df_ret.copy()
    df_ret["Tempo_min"] = None

    for idx, row_ret in df_ret.iterrows():
        maq   = row_ret.get(maq_col_ret)
        dt_ret = row_ret.get(col_data_ret)
        if pd.isna(maq) or pd.isna(dt_ret):
            continue

        # Chamados da mesma máquina, anteriores ao retorno
        mask = (
            (df_ch[col_maquina] == maq) &
            (df_ch[col_data_cham] <= dt_ret)
        )
        candidatos = df_ch.loc[mask, col_data_cham].dropna()
        if candidatos.empty:
            continue

        # Pega o mais próximo (mais recente)
        dt_cham = candidatos.max()
        delta_min = (dt_ret - dt_cham).total_seconds() / 60

        # Descarta valores negativos ou absurdos (> 30 dias)
        if 0 < delta_min <= 43200:
            df_ret.at[idx, "Tempo_min"] = delta_min

    df_ret["Tempo_min"] = pd.to_numeric(df_ret["Tempo_min"], errors="coerce")
    return df_ret


# ── Manutenção Preventiva ──────────────────────────────────────────────────────
@st.cache_data(ttl=300)
def load_preventiva(sheet_id: str, gid: str = "0") -> pd.DataFrame:
    """
    Lê a planilha de PMs preventivas.
    Colunas esperadas: Patrimônio | Tarefa | Tipo | Frequência_Dias |
                       Última_Execução | Responsável | Prioridade | Observações
    """
    if not sheet_id.strip():
        return pd.DataFrame()
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    try:
        resp = requests.get(url, timeout=15)
        resp.raise_for_status()
        df = pd.read_csv(StringIO(resp.text))
        df.columns = df.columns.str.strip()
        return df
    except Exception as e:
        st.warning(f"Planilha preventiva: {e}")
        return pd.DataFrame()


def calcular_status_pm(row, antecedencia: int = 7) -> tuple[str, str]:
    """Retorna (status_texto, cor_hex) baseado em dias restantes."""
    dias = row.get("Dias_Restantes")
    if pd.isna(dias):
        return "Sem data", "#888888"
    dias = int(dias)
    if dias < 0:
        return f"⛔ Atrasada {abs(dias)}d", "#e74c3c"
    elif dias == 0:
        return "🚨 Vence hoje!", "#e74c3c"
    elif dias <= antecedencia:
        return f"⚠️ Em {dias}d", "#f39c12"
    elif dias <= 30:
        return f"🔔 Em {dias}d", "#3498db"
    else:
        return f"✅ Em {dias}d", "#2ecc71"


def calcular_mtbf_mttr(df_ch, df_ret, col_data_cham, col_maquina, col_data_ret, maq_col_ret):
    """Calcula MTBF e MTTR por máquina."""
    resultado = {}

    # MTTR — já temos Tempo_min em df_ret
    if not df_ret.empty and maq_col_ret and "Tempo_min" in df_ret.columns:
        mttr = df_ret.groupby(maq_col_ret)["Tempo_min"].mean()
        for maq, val in mttr.items():
            resultado.setdefault(maq, {})["MTTR_min"] = val

    # MTBF — tempo médio entre falhas consecutivas por máquina
    if (not df_ch.empty and col_data_cham != "(nenhuma)"
            and col_maquina != "(nenhuma)"):
        for maq, grp in df_ch.groupby(col_maquina):
            datas = grp[col_data_cham].dropna().sort_values()
            if len(datas) >= 2:
                deltas = datas.diff().dropna().dt.total_seconds() / 3600  # horas
                mtbf_h = deltas.mean()
                resultado.setdefault(maq, {})["MTBF_h"] = mtbf_h

    if not resultado:
        return pd.DataFrame()

    df_res = pd.DataFrame(resultado).T.reset_index()
    df_res.columns = ["Máquina"] + [c for c in df_res.columns if c != "index"]
    return df_res


def gerar_msg_telegram(row, antecedencia: int) -> str:
    maq   = row.get("Patrimônio", "?")
    nome  = catalogo.get(str(maq).upper(), {}).get("nome", "")
    tarefa = row.get("Tarefa", "")
    data  = row.get("Próxima_Execução", "")
    resp  = row.get("Responsável", "")
    prior = row.get("Prioridade", "")
    dias  = int(row.get("Dias_Restantes", 0))

    if isinstance(data, pd.Timestamp):
        data = data.strftime("%d/%m/%Y")

    emoji_prior = {"Alta": "🔴", "Média": "🟡", "Baixa": "🟢"}.get(str(prior), "⚙️")
    aviso = "⛔ *ATRASADA*" if dias < 0 else f"🔔 Vence em *{dias} dias*"

    return (
        f"🔧 *MANUTENÇÃO PREVENTIVA*\n\n"
        f"🏭 Máquina: *{maq}*{(' – ' + nome.title()) if nome else ''}\n"
        f"📋 Tarefa: {tarefa}\n"
        f"📅 Data prevista: *{data}*\n"
        f"👷 Responsável: {resp}\n"
        f"{emoji_prior} Prioridade: {prior}\n"
        f"{aviso}\n\n"
        f"_Agende a manutenção com antecedência de {antecedencia} dias._"
    )


def _gerar_importacao_historico(df_ch, col_tipo, col_maq, col_prob, col_data, col_mec, catalogo):
    """
    Filtra os chamados do tipo 'preventiva' e gera uma tabela
    pronta para copiar/colar na planilha de PMs.
    """
    st.subheader("📥 Importar Preventivas do Histórico de Chamados")

    if df_ch.empty or col_tipo == "(nenhuma)" or col_tipo not in df_ch.columns:
        st.warning("Configure a coluna **'Tipo Manutenção'** nos chamados (menu lateral → Colunas – Chamados) para usar esta função.")
        return

    # Filtra preventivas
    mask_prev = df_ch[col_tipo].astype(str).str.lower().str.contains("prev", na=False)
    df_prev_hist = df_ch[mask_prev].copy()

    if df_prev_hist.empty:
        st.info("Nenhum chamado do tipo 'preventiva' encontrado no histórico.")
        return

    st.success(f"✅ {len(df_prev_hist)} chamados preventivos encontrados no histórico.")

    # Para cada máquina, pega o mais recente
    rows_out = []
    col_maq_ok  = col_maq  if col_maq  != "(nenhuma)" else None
    col_prob_ok = col_prob if col_prob != "(nenhuma)" else None
    col_data_ok = col_data if col_data != "(nenhuma)" else None
    col_mec_ok  = col_mec  if col_mec  != "(nenhuma)" else None

    grp_col = col_maq_ok or df_prev_hist.columns[0]

    for maquina, grp in df_prev_hist.groupby(grp_col):
        # Data mais recente
        ultima = None
        if col_data_ok and col_data_ok in grp.columns:
            datas = pd.to_datetime(grp[col_data_ok], errors="coerce").dropna()
            if not datas.empty:
                ultima = datas.max()

        # Descrição mais comum (tarefa)
        tarefa = ""
        if col_prob_ok and col_prob_ok in grp.columns:
            tarefa = grp[col_prob_ok].dropna().mode()
            tarefa = str(tarefa.iloc[0]) if len(tarefa) > 0 else ""
            tarefa = tarefa[:80]  # limita tamanho

        # Mecânico mais frequente
        responsavel = ""
        if col_mec_ok and col_mec_ok in grp.columns:
            resp = grp[col_mec_ok].dropna().mode()
            responsavel = str(resp.iloc[0]) if len(resp) > 0 else ""

        # Nome da máquina no catálogo
        nome_maq = catalogo.get(str(maquina).upper(), {}).get("nome", "")

        rows_out.append({
            "Patrimônio"     : str(maquina),
            "Nome Máquina"   : nome_maq.title() if nome_maq else "",
            "Tarefa"         : tarefa,
            "Tipo"           : "Inspeção",          # sugestão padrão — edite conforme necessário
            "Frequência_Dias": 30,                   # sugestão padrão — edite conforme necessário
            "Última_Execução": ultima.strftime("%d/%m/%Y") if ultima else "",
            "Responsável"    : responsavel,
            "Prioridade"     : "Média",              # sugestão padrão
            "Observações"    : f"Importado do histórico ({len(grp)} registros)"
        })

    df_out = pd.DataFrame(rows_out).sort_values("Patrimônio")

    st.markdown("#### Tabela para copiar na planilha de PMs")
    st.caption("Revise os campos **Tarefa**, **Tipo**, **Frequência_Dias** e **Prioridade** antes de colar — foram preenchidos com sugestões padrão.")
    st.dataframe(df_out, use_container_width=True, hide_index=True)

    # Download CSV (abre no Excel e copia facilmente)
    csv_bytes = df_out.to_csv(index=False, sep="\t", encoding="utf-8-sig").encode("utf-8-sig")
    st.download_button(
        "⬇ Baixar como planilha (TSV — abre no Excel)",
        data=csv_bytes,
        file_name="preventivas_para_importar.csv",
        mime="text/tab-separated-values",
        use_container_width=True
    )

    st.info("""
    **Como usar:**
    1. Clique em **Baixar** → abre no Excel
    2. Revise e ajuste **Tarefa**, **Tipo**, **Frequência** e **Prioridade** de cada máquina
    3. Copie as linhas e cole na sua **planilha de PMs preventivas**
    4. Cole o ID dessa planilha no menu lateral do dashboard
    """)


def calendario_mensal(df_pm: pd.DataFrame, ano: int, mes: int) -> go.Figure:
    """Gera um heatmap-calendário do mês com PMs."""
    import calendar
    cal = calendar.monthcalendar(ano, mes)
    dias_semana = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]

    # Conta PMs por dia
    pms_dia: dict[int, list] = {}
    if not df_pm.empty and "Próxima_Execução" in df_pm.columns:
        for _, row in df_pm.iterrows():
            dt = row["Próxima_Execução"]
            if pd.notna(dt) and dt.year == ano and dt.month == mes:
                d = dt.day
                pms_dia.setdefault(d, []).append(
                    row.get("Patrimônio","") + " " + row.get("Tarefa","")
                )

    # Monta matriz
    z, text, hover = [], [], []
    for semana in cal:
        z_row, t_row, h_row = [], [], []
        for dia in semana:
            if dia == 0:
                z_row.append(None); t_row.append(""); h_row.append("")
            else:
                qtd = len(pms_dia.get(dia, []))
                z_row.append(qtd)
                t_row.append(f"<b>{dia}</b>" + (f"<br>{qtd} PM" if qtd else ""))
                detalhe = "<br>".join(pms_dia.get(dia, []))
                h_row.append(f"Dia {dia}<br>{detalhe}" if detalhe else f"Dia {dia} — sem PM")
        z.append(z_row); text.append(t_row); hover.append(h_row)

    fig = go.Figure(go.Heatmap(
        z=z, text=text, hovertext=hover,
        texttemplate="%{text}", hovertemplate="%{hovertext}<extra></extra>",
        colorscale=[[0,"#1e2130"],[0.01,"#2ecc71"],[0.5,"#f39c12"],[1,"#e74c3c"]],
        showscale=False, xgap=3, ygap=3,
    ))
    fig.update_xaxes(tickvals=list(range(7)), ticktext=dias_semana, side="top")
    fig.update_yaxes(visible=False, autorange="reversed")
    fig.update_layout(
        title=f"📅 {calendar.month_name[mes]} {ano}",
        height=320, margin=dict(t=60, b=10, l=10, r=10),
        paper_bgcolor="#0e1117", plot_bgcolor="#0e1117",
        font=dict(color="#fafafa"),
    )
    return fig


# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style='text-align:center; padding:8px 0 4px 0'>
      <span style='font-size:2.2rem'>🏭</span><br>
      <span style='font-size:1.3rem; font-weight:800; color:#F97316; letter-spacing:1px'>MASSAFEST</span><br>
      <span style='font-size:0.75rem; color:#888; letter-spacing:2px'>PCM · MANUTENÇÃO</span>
    </div>
    """, unsafe_allow_html=True)
    st.title("")
    st.markdown("---")
    st.subheader("Planilhas – Chamados/Retornos")
    sheet_chamados = st.text_input("ID – Chamados", DEFAULT_SHEET_CHAMADOS)
    gid_chamados   = st.text_input("GID – Chamados", DEFAULT_GID_CHAMADOS)
    sheet_retornos = st.text_input("ID – Retornos", DEFAULT_SHEET_RETORNOS)
    gid_retornos   = st.text_input("GID – Retornos", DEFAULT_GID_RETORNOS)
    st.subheader("Catálogo de Máquinas")
    sheet_cat1 = st.text_input("ID – Catálogo 1 (MP)", DEFAULT_SHEET_CAT1)
    sheet_cat2 = st.text_input("ID – Catálogo 2 (PA)", DEFAULT_SHEET_CAT2)
    st.subheader("Manutenção Preventiva")
    sheet_preventiva = st.text_input("ID – Planilha de PMs", DEFAULT_SHEET_PREVENTIVA,
                                     placeholder="Cole o ID da planilha aqui")
    gid_preventiva   = st.text_input("GID – Aba de PMs", "0")
    antecedencia     = st.slider("⏰ Alertar com antecedência (dias)", 1, 30, 7)
    if st.button("🔄 Atualizar dados"):
        st.cache_data.clear()
    st.markdown("---")
    st.subheader("Agrupamento de nomes")
    similaridade = st.slider(
        "Sensibilidade (nomes parecidos)",
        min_value=0.50, max_value=1.00, value=0.80, step=0.05,
        help="Quanto menor, mais nomes são agrupados. 0.80 é o recomendado."
    )
    st.caption("Agrupa automaticamente variações como 'Alexandre' e 'Alexadre'.")
    st.markdown("---")
    st.caption("Dados atualizados a cada 5 min automaticamente.")

with st.spinner("Carregando planilhas…"):
    df_chamados = load_sheet(sheet_chamados, gid_chamados)
    df_retornos = load_sheet(sheet_retornos, gid_retornos)

with st.spinner("Carregando catálogo de máquinas…"):
    cat1, hist_leg1 = load_catalogo(sheet_cat1)
    cat2, hist_leg2 = load_catalogo(sheet_cat2)

# Catálogo unificado: {codigo → {nome, localizacao}}
catalogo = {**cat1, **cat2}

# Histórico legado unificado
hist_legado = pd.concat([hist_leg1, hist_leg2], ignore_index=True) if (not hist_leg1.empty or not hist_leg2.empty) else pd.DataFrame()

# Mapa de conversão: número legado → código MP/PA
# Ex: '003' → 'MP003', '045' → 'PA045'
legacy_map = {}
for cod in catalogo:
    num = re.sub(r'[^0-9]', '', cod).lstrip('0')
    if num:
        legacy_map[num.zfill(3)] = cod
        legacy_map[num]          = cod

if df_chamados.empty and df_retornos.empty:
    st.warning("Nenhum dado encontrado. Verifique se as planilhas estão públicas.")
    st.stop()

st.sidebar.markdown("---")
st.sidebar.subheader("Colunas – Chamados")

def col_select(label, df, default_hints, key):
    cols = ["(nenhuma)"] + list(df.columns)
    best = next((c for h in default_hints for c in df.columns if h.lower() in c.lower()), "(nenhuma)")
    idx = cols.index(best) if best in cols else 0
    return st.sidebar.selectbox(label, cols, index=idx, key=key)

if not df_chamados.empty:
    col_data_cham   = col_select("Data/Hora",   df_chamados, ["timestamp","data","hora","abertura"], "c_data")
    col_maquina     = col_select("Máquina",     df_chamados, ["maquina","máquina","equipamento","ativo"], "c_maq")
    col_patrimonio  = col_select("Patrimônio",  df_chamados, ["patrimonio","patrimônio","tag","codigo"], "c_pat")
    col_problema    = col_select("Problema",    df_chamados, ["problema","descricao","descri","falha","defeito"], "c_prob")
    col_encarregado = col_select("Encarregado", df_chamados, ["encarregado","solicitante","abertura","nome"], "c_enc")
    col_tipo_manut  = col_select("Tipo Manutenção", df_chamados, ["tipo","manutencao","manutenção","corretiva","preventiva"], "c_tipo")
else:
    col_data_cham = col_maquina = col_patrimonio = col_problema = col_encarregado = col_tipo_manut = "(nenhuma)"

st.sidebar.subheader("Colunas – Retornos")
if not df_retornos.empty:
    col_data_ret = col_select("Data/Hora", df_retornos, ["timestamp","data","hora","conclusao"], "r_data")
    col_mecanico = col_select("Mecânico",  df_retornos, ["mecanico","mecânico","tecnico","técnico","executor"], "r_mec")
    col_maq_ret  = col_select("Máquina",   df_retornos, ["maquina","máquina","equipamento","ativo"], "r_maq")
    col_servico  = col_select("Serviço",   df_retornos, ["servico","serviço","atividade","descri","relatorio"], "r_serv")
    col_status   = col_select("Status",    df_retornos, ["status","situacao","situação","concluido"], "r_stat")
else:
    col_data_ret = col_mecanico = col_maq_ret = col_servico = col_status = "(nenhuma)"

# ── Pré-processar ──────────────────────────────────────────────────────────────
if not df_chamados.empty and col_data_cham != "(nenhuma)":
    df_chamados = parse_date_col(df_chamados, col_data_cham)
if not df_retornos.empty and col_data_ret != "(nenhuma)":
    df_retornos = parse_date_col(df_retornos, col_data_ret)
if not df_retornos.empty and col_mecanico != "(nenhuma)" and col_mecanico in df_retornos.columns:
    # 1. Expande duplas/trios → cada mecânico vira uma linha separada
    df_retornos = expandir_mecanicos(df_retornos, col_mecanico)
    # 2. Agrupa nomes similares (erros de digitação)
    df_retornos[col_mecanico] = agrupar_nomes_similares(df_retornos[col_mecanico], threshold=similaridade)

maq_col_ret = col_maq_ret if col_maq_ret != "(nenhuma)" else None

# ── Normalizar máquinas ────────────────────────────────────────────────────────
for _df, _col in [(df_chamados, col_maquina), (df_retornos, maq_col_ret)]:
    if _df is not None and not _df.empty and _col and _col != "(nenhuma)" and _col in _df.columns:
        _df[_col] = _df[_col].apply(normalizar_maquina)
        # Remove linhas com máquina inválida (campo vazio após normalização)
        mask_valido = _df[_col].str.strip().astype(bool)
        _df.drop(_df[~mask_valido].index, inplace=True)
        _df.reset_index(drop=True, inplace=True)

# ── Calcular tempos ────────────────────────────────────────────────────────────
df_retornos = calcular_tempos(df_chamados, df_retornos, col_data_cham, col_maquina, col_data_ret, maq_col_ret)

# ── Carregar e processar preventiva ───────────────────────────────────────────
df_prev_raw = load_preventiva(sheet_preventiva, gid_preventiva)
df_prev     = pd.DataFrame()

if not df_prev_raw.empty:
    df_prev = df_prev_raw.copy()
    # Detecta colunas flexíveis
    def _find_col(df, hints):
        for h in hints:
            for c in df.columns:
                if h.lower() in c.lower():
                    return c
        return None

    c_pat  = _find_col(df_prev, ["patrimônio","patrimonio","maquina","máquina","cod"])
    c_tar  = _find_col(df_prev, ["tarefa","task","serv","atividade"])
    c_tipo = _find_col(df_prev, ["tipo","type"])
    c_freq = _find_col(df_prev, ["frequência","frequencia","dias","freq"])
    c_ult  = _find_col(df_prev, ["última","ultima","last","execu"])
    c_resp = _find_col(df_prev, ["responsável","responsavel","mecânico","mecanico"])
    c_pri  = _find_col(df_prev, ["prioridade","prior","urgent"])
    c_obs  = _find_col(df_prev, ["observ","nota","obs"])

    rename = {}
    for orig, novo in [(c_pat,"Patrimônio"),(c_tar,"Tarefa"),(c_tipo,"Tipo"),
                       (c_freq,"Frequência_Dias"),(c_ult,"Última_Execução"),
                       (c_resp,"Responsável"),(c_pri,"Prioridade"),(c_obs,"Observações")]:
        if orig and orig not in rename.values():
            rename[orig] = novo
    df_prev = df_prev.rename(columns=rename)

    if "Última_Execução" in df_prev.columns:
        df_prev["Última_Execução"] = pd.to_datetime(df_prev["Última_Execução"], dayfirst=True, errors="coerce")
    if "Frequência_Dias" in df_prev.columns:
        df_prev["Frequência_Dias"] = pd.to_numeric(df_prev["Frequência_Dias"], errors="coerce")
    if "Última_Execução" in df_prev.columns and "Frequência_Dias" in df_prev.columns:
        df_prev["Próxima_Execução"] = df_prev["Última_Execução"] + pd.to_timedelta(df_prev["Frequência_Dias"], unit="D")
        df_prev["Dias_Restantes"]   = (df_prev["Próxima_Execução"] - pd.Timestamp.today()).dt.days

    df_prev["Status"], df_prev["Cor"] = zip(*df_prev.apply(
        lambda r: calcular_status_pm(r, antecedencia), axis=1
    )) if len(df_prev) else ([], [])

# MTBF/MTTR
df_mtbf = calcular_mtbf_mttr(df_chamados, df_retornos, col_data_cham, col_maquina, col_data_ret, maq_col_ret)

# ── Filtros principais ─────────────────────────────────────────────────────────
st.sidebar.markdown("---")
st.sidebar.subheader("Filtros")

all_dates = []
if not df_chamados.empty and col_data_cham != "(nenhuma)":
    all_dates += list(df_chamados[col_data_cham].dropna())
if not df_retornos.empty and col_data_ret != "(nenhuma)":
    all_dates += list(df_retornos[col_data_ret].dropna())

if all_dates:
    min_date = min(all_dates).date()
    max_date = max(all_dates).date()
else:
    min_date = datetime.today().date() - timedelta(days=365)
    max_date = datetime.today().date()

periodo = st.sidebar.date_input("Período", value=(min_date, max_date), min_value=min_date, max_value=max_date)
if len(periodo) == 2:
    dt_ini, dt_fim = pd.Timestamp(periodo[0]), pd.Timestamp(periodo[1]) + timedelta(days=1)
else:
    dt_ini, dt_fim = pd.Timestamp(min_date), pd.Timestamp(max_date) + timedelta(days=1)

mecánicos = []
if not df_retornos.empty and col_mecanico != "(nenhuma)":
    mecánicos = sorted(df_retornos[col_mecanico].dropna().unique().tolist())
sel_mecanicos = st.sidebar.multiselect("Mecânicos", mecánicos, default=mecánicos)

maquinas = []
if not df_chamados.empty and col_maquina != "(nenhuma)":
    maquinas += df_chamados[col_maquina].dropna().unique().tolist()
if not df_retornos.empty and maq_col_ret:
    maquinas += df_retornos[maq_col_ret].dropna().unique().tolist()
maquinas = sorted(set(maquinas))

# Filtro por linha de produção
import re as _re
def _linha(m):
    if _re.match(r'(?i)^MP', str(m)): return "🥟 Pastel (MP)"
    if _re.match(r'(?i)^PA', str(m)): return "🧄 Pão de Alho (PA)"
    if str(m) == "S/N":               return "❓ Sem número"
    return "📦 Legado (número)"

linhas_disp = sorted(set(_linha(m) for m in maquinas))
sel_linhas  = st.sidebar.multiselect("Linha de produção", linhas_disp, default=linhas_disp)
maquinas    = [m for m in maquinas if _linha(m) in sel_linhas]
sel_maquinas = st.sidebar.multiselect("Máquinas", maquinas, default=maquinas)

def filter_df(df, date_col, maq_col_name, mec_col_name=None):
    if df.empty:
        return df
    if date_col != "(nenhuma)" and date_col in df.columns:
        df = df[df[date_col].between(dt_ini, dt_fim, inclusive="left")]
    if maq_col_name and maq_col_name in df.columns and sel_maquinas:
        df = df[df[maq_col_name].isin(sel_maquinas)]
    if mec_col_name and mec_col_name in df.columns and sel_mecanicos:
        df = df[df[mec_col_name].isin(sel_mecanicos)]
    return df

df_ch  = filter_df(df_chamados, col_data_cham, col_maquina if col_maquina != "(nenhuma)" else None)
df_ret = filter_df(df_retornos, col_data_ret,  maq_col_ret, col_mecanico if col_mecanico != "(nenhuma)" else None)

# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div style='padding:18px 0 8px 0'>
  <span style='font-size:2rem; font-weight:900; color:#F97316'>🏭 MASSAFEST</span>
  <span style='font-size:1.3rem; font-weight:400; color:#aaa; margin-left:12px'>Painel de Manutenção – PCM</span>
</div>
""", unsafe_allow_html=True)

def nome_maquina(codigo: str) -> str:
    """Retorna 'MP003 – MÁQUINA DE CORTE' ou só o código se não encontrar."""
    info = catalogo.get(str(codigo).strip().upper())
    if info and info.get("nome"):
        return f"{codigo} – {info['nome'].title()}"
    return str(codigo)

tab_visao, tab_mecanico, tab_maquina, tab_historico, tab_prev, tab_relatorio, tab_dados = st.tabs([
    "📊 Visão Geral", "👷 Por Mecânico", "🏭 Por Máquina", "📋 Histórico",
    "🔧 Preventiva", "📄 Relatórios", "📂 Dados Brutos"
])

# ─────────────────────────────────────────────
# ABA 1: Visão Geral
# ─────────────────────────────────────────────
with tab_visao:
    c1, c2, c3, c4, c5 = st.columns(5)
    total_chamados  = len(df_ch)
    total_retornos  = len(df_ret)
    total_maquinas  = len(set(
        (list(df_ch[col_maquina].dropna()) if col_maquina != "(nenhuma)" and not df_ch.empty else []) +
        (list(df_ret[maq_col_ret].dropna()) if maq_col_ret and not df_ret.empty else [])
    ))
    total_mecanicos = df_ret[col_mecanico].nunique() if col_mecanico != "(nenhuma)" and not df_ret.empty else 0

    tempo_medio_geral = None
    if "Tempo_min" in df_ret.columns:
        tempo_medio_geral = df_ret["Tempo_min"].dropna().mean()

    c1.metric("Total de Chamados", total_chamados)
    c2.metric("Atendimentos Realizados", total_retornos)
    c3.metric("Máquinas Atendidas", total_maquinas)
    c4.metric("Mecânicos Ativos", total_mecanicos)
    c5.metric("Tempo Médio de Atendimento",
              formatar_tempo(tempo_medio_geral) if tempo_medio_geral else "—")

    st.markdown("---")
    col_l, col_r = st.columns(2)

    with col_l:
        st.subheader("Chamados ao longo do tempo")
        if not df_ch.empty and col_data_cham != "(nenhuma)":
            granularidade = st.selectbox("Granularidade", ["Dia", "Semana", "Mês"], key="gran_visao")
            freq_map = {"Dia": "D", "Semana": "W", "Mês": "ME"}
            ts = df_ch.set_index(col_data_cham).resample(freq_map[granularidade]).size().reset_index(name="Chamados")
            fig = px.bar(ts, x=col_data_cham, y="Chamados", color_discrete_sequence=[COR_PRIMARIA])
            fig.update_layout(margin=dict(t=10, b=10))
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Configure a coluna de data dos chamados no menu lateral.")

    with col_r:
        st.subheader("Atendimentos por mecânico")
        if not df_ret.empty and col_mecanico != "(nenhuma)":
            mec_count = df_ret[col_mecanico].value_counts().reset_index()
            mec_count.columns = ["Mecânico", "Atendimentos"]
            fig = px.bar(mec_count, x="Atendimentos", y="Mecânico", orientation="h",
                         color="Atendimentos", color_continuous_scale="Oranges")
            fig.update_layout(margin=dict(t=10, b=10), yaxis={"categoryorder": "total ascending"})
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Configure a coluna de mecânico no menu lateral.")

# ─────────────────────────────────────────────
# ABA 2: Por Mecânico
# ─────────────────────────────────────────────
with tab_mecanico:
    st.subheader("Desempenho por Mecânico")
    if df_ret.empty or col_mecanico == "(nenhuma)":
        st.info("Dados de retorno não disponíveis ou coluna de mecânico não configurada.")
    else:
        mec_df = df_ret[col_mecanico].value_counts().reset_index()
        mec_df.columns = ["Mecânico", "Atendimentos"]
        mec_df["% do Total"] = (mec_df["Atendimentos"] / mec_df["Atendimentos"].sum() * 100).round(1)

        # Adiciona tempo médio por mecânico se disponível
        if "Tempo_min" in df_ret.columns:
            tempo_mec = df_ret.groupby(col_mecanico)["Tempo_min"].mean().reset_index()
            tempo_mec.columns = ["Mecânico", "Tempo_Médio_min"]
            mec_df = mec_df.merge(tempo_mec, on="Mecânico", how="left")
            mec_df["Tempo Médio"] = mec_df["Tempo_Médio_min"].apply(formatar_tempo)
            mec_df = mec_df.drop(columns=["Tempo_Médio_min"])

        col_a, col_b = st.columns([1, 2])
        with col_a:
            st.dataframe(mec_df, use_container_width=True, hide_index=True)
        with col_b:
            fig = px.pie(mec_df, names="Mecânico", values="Atendimentos",
                         title="Distribuição de atendimentos",
                         color_discrete_sequence=px.colors.qualitative.Set2)
            fig.update_traces(textposition="inside", textinfo="percent+label")
            st.plotly_chart(fig, use_container_width=True)

        if col_data_ret != "(nenhuma)":
            st.markdown("---")
            st.subheader("Evolução mensal por mecânico")
            df_ret_copy = df_ret.copy()
            df_ret_copy["Mês"] = df_ret_copy[col_data_ret].dt.to_period("M").astype(str)
            mensal = df_ret_copy.groupby(["Mês", col_mecanico]).size().reset_index(name="Atendimentos")
            fig2 = px.line(mensal, x="Mês", y="Atendimentos", color=col_mecanico, markers=True)
            fig2.update_layout(xaxis_tickangle=-45)
            st.plotly_chart(fig2, use_container_width=True)

# ─────────────────────────────────────────────
# ABA 3: Por Máquina
# ─────────────────────────────────────────────
with tab_maquina:
    st.subheader("Máquinas com mais manutenção")
    maq_src_col = col_maquina if (col_maquina != "(nenhuma)" and not df_ch.empty) else None
    maq_ret_col = maq_col_ret if (maq_col_ret and not df_ret.empty) else None

    frames = []
    if maq_src_col:
        c = df_ch[maq_src_col].value_counts().reset_index()
        c.columns = ["Máquina", "Chamados"]
        frames.append(c.set_index("Máquina"))
    if maq_ret_col:
        c = df_ret[maq_ret_col].value_counts().reset_index()
        c.columns = ["Máquina", "Atendimentos"]
        frames.append(c.set_index("Máquina"))

    # Tempo de manutenção por máquina
    if maq_ret_col and "Tempo_min" in df_ret.columns:
        t = df_ret.groupby(maq_ret_col)["Tempo_min"].agg(
            Tempo_Médio_min="mean",
            Tempo_Total_min="sum",
            Tempo_Máximo_min="max"
        ).reset_index()
        t.columns = ["Máquina", "Tempo_Médio_min", "Tempo_Total_min", "Tempo_Máximo_min"]
        frames.append(t.set_index("Máquina"))

    if frames:
        maq_df = pd.concat(frames, axis=1).reset_index()

        # Formata colunas de tempo
        for col_t in ["Tempo_Médio_min", "Tempo_Total_min", "Tempo_Máximo_min"]:
            if col_t in maq_df.columns:
                label = col_t.replace("_min","").replace("_"," ")
                maq_df[label] = maq_df[col_t].apply(formatar_tempo)
                maq_df = maq_df.drop(columns=[col_t])

        # Coluna total de ocorrências
        num_cols = [c for c in ["Chamados","Atendimentos"] if c in maq_df.columns]
        if num_cols:
            maq_df["Total"] = maq_df[num_cols].apply(pd.to_numeric, errors="coerce").sum(axis=1)
            maq_df = maq_df.sort_values("Total", ascending=False)

        # Coluna de linha de produção
        maq_df.insert(1, "Linha", maq_df["Máquina"].apply(_linha))

        col_t, col_g = st.columns([1, 2])
        with col_t:
            st.dataframe(maq_df.fillna("—"), use_container_width=True, hide_index=True)
        with col_g:
            if "Total" in maq_df.columns:
                fig = px.bar(maq_df.head(20), x="Máquina", y="Total",
                             title="Top 20 máquinas por ocorrências",
                             color="Total", color_continuous_scale=ESCALA_CALOR)
                fig.update_layout(xaxis_tickangle=-45)
                st.plotly_chart(fig, use_container_width=True)

        # Gráfico de tempo médio por máquina
        if "Tempo Médio" in maq_df.columns and maq_ret_col and "Tempo_min" in df_ret.columns:
            st.markdown("---")
            st.subheader("⏱ Tempo médio de atendimento por máquina")
            tempo_graf = df_ret.groupby(maq_ret_col)["Tempo_min"].mean().reset_index()
            tempo_graf.columns = ["Máquina", "Minutos"]
            tempo_graf = tempo_graf.dropna().sort_values("Minutos", ascending=False).head(20)
            tempo_graf["Tempo"] = tempo_graf["Minutos"].apply(formatar_tempo)
            fig3 = px.bar(tempo_graf, x="Máquina", y="Minutos",
                          text="Tempo",
                          title="Top 20 – Maior tempo médio de atendimento",
                          color="Minutos", color_continuous_scale="Oranges")
            fig3.update_traces(textposition="outside")
            fig3.update_layout(xaxis_tickangle=-45, yaxis_title="Minutos")
            st.plotly_chart(fig3, use_container_width=True)
    else:
        st.info("Configure as colunas de máquina no menu lateral.")

# ─────────────────────────────────────────────
# ABA 4: Histórico por Máquina
# ─────────────────────────────────────────────
with tab_historico:
    st.subheader("Histórico de manutenção por máquina")
    maquinas_disp = []
    if maq_src_col:
        maquinas_disp += df_ch[maq_src_col].dropna().unique().tolist()
    if maq_ret_col:
        maquinas_disp += df_ret[maq_ret_col].dropna().unique().tolist()
    maquinas_disp = sorted(set(maquinas_disp))

    if not maquinas_disp:
        st.info("Nenhuma máquina encontrada. Configure as colunas no menu lateral.")
    else:
        maq_sel = st.selectbox("Selecione a máquina", maquinas_disp)
        hist_rows = []

        if maq_src_col and not df_ch.empty:
            subset = df_ch[df_ch[maq_src_col] == maq_sel].copy()
            subset["Origem"] = "Chamado"
            rename = {}
            if col_data_cham   != "(nenhuma)": rename[col_data_cham]   = "Data/Hora"
            if col_problema    != "(nenhuma)": rename[col_problema]    = "Descrição"
            if col_encarregado != "(nenhuma)": rename[col_encarregado] = "Encarregado"
            subset = subset.rename(columns=rename)
            hist_rows.append(subset)

        if maq_ret_col and not df_ret.empty:
            subset = df_ret[df_ret[maq_ret_col] == maq_sel].copy()
            subset["Origem"] = "Atendimento"
            rename = {}
            if col_data_ret != "(nenhuma)": rename[col_data_ret] = "Data/Hora"
            if col_servico  != "(nenhuma)": rename[col_servico]  = "Descrição"
            if col_mecanico != "(nenhuma)": rename[col_mecanico] = "Mecânico"
            if col_status   != "(nenhuma)": rename[col_status]   = "Status"
            subset = subset.rename(columns=rename)
            if "Tempo_min" in subset.columns:
                subset["Tempo de Atendimento"] = subset["Tempo_min"].apply(formatar_tempo)
                subset = subset.drop(columns=["Tempo_min"])
            hist_rows.append(subset)

        # Inclui histórico legado da planilha de cadastro
        if not hist_legado.empty and "Máquina" in hist_legado.columns:
            leg_maq = hist_legado[hist_legado["Máquina"] == maq_sel].copy()
            if not leg_maq.empty:
                hist_rows.append(leg_maq)

        if hist_rows:
            hist = pd.concat(hist_rows, ignore_index=True)
            if "Data/Hora" in hist.columns:
                hist = hist.sort_values("Data/Hora", ascending=False)

            cols_show = [c for c in ["Data/Hora","Origem","Tipo","Descrição","Troca de Peça","Mecânico","Tempo de Atendimento","Encarregado","Status","Custo"] if c in hist.columns]
            st.dataframe(hist[cols_show], use_container_width=True, hide_index=True)

            # Métricas rápidas
            m1, m2, m3 = st.columns(3)
            m1.metric("Total de ocorrências", len(hist))
            if "Mecânico" in hist.columns:
                top_mec = hist["Mecânico"].value_counts().idxmax() if not hist["Mecânico"].isna().all() else "—"
                m2.metric("Mecânico que mais atendeu", top_mec)
            if "Tempo de Atendimento" in hist.columns and maq_ret_col and "Tempo_min" in df_ret.columns:
                subset_t = df_ret[df_ret[maq_ret_col] == maq_sel]["Tempo_min"].dropna()
                if not subset_t.empty:
                    m3.metric("Tempo médio nesta máquina", formatar_tempo(subset_t.mean()))
        else:
            st.info("Nenhum registro encontrado para esta máquina no período selecionado.")

# ─────────────────────────────────────────────
# ─────────────────────────────────────────────
# ABA 5: Manutenção Preventiva
# ─────────────────────────────────────────────
with tab_prev:
    st.title("🔧 Gestão de Manutenção Preventiva")

    # ── KPIs ──────────────────────────────────────────────────────────────────
    k1, k2, k3, k4, k5 = st.columns(5)

    if df_prev.empty:
        k1.metric("PMs Cadastradas", 0)
        k2.metric("⛔ Atrasadas", 0)
        k3.metric("⚠️ Próximas 7 dias", 0)
        k4.metric("Taxa Prev/Corr", "—")
        k5.metric("MTTR Médio", "—")
        st.info("""
        ### Como configurar a Manutenção Preventiva
        **1. Crie uma nova planilha Google Sheets** com as colunas abaixo:

        | Patrimônio | Tarefa | Tipo | Frequência_Dias | Última_Execução | Responsável | Prioridade | Observações |
        |---|---|---|---|---|---|---|---|
        | MP003 | Lubrificação correntes | Lubrificação | 30 | 01/05/2026 | Fernando | Alta | |

        **2. Compartilhe a planilha** como pública · **3. Cole o ID** no campo do menu lateral

        **Tipos sugeridos:** Lubrificação · Inspeção · Limpeza · Calibração · Troca de filtro · Revisão geral
        """)

        # ── Importar preventivos do histórico ─────────────────────────────
        st.markdown("---")
        _gerar_importacao_historico(df_ch, col_tipo_manut, col_maquina, col_problema,
                                    col_data_cham, col_mecanico, catalogo)
    else:
        total_pms    = len(df_prev)
        atrasadas    = (df_prev["Dias_Restantes"] < 0).sum() if "Dias_Restantes" in df_prev.columns else 0
        proximas_7   = ((df_prev["Dias_Restantes"] >= 0) & (df_prev["Dias_Restantes"] <= 7)).sum() if "Dias_Restantes" in df_prev.columns else 0
        proximas_30  = ((df_prev["Dias_Restantes"] >= 0) & (df_prev["Dias_Restantes"] <= 30)).sum() if "Dias_Restantes" in df_prev.columns else 0

        # Taxa preventiva vs corretiva
        n_prev = 0; n_corr = 0
        if not df_ch.empty:
            for col in df_ch.columns:
                if "tipo" in col.lower() or "manut" in col.lower():
                    n_prev = df_ch[col].str.lower().str.contains("prev", na=False).sum()
                    n_corr = df_ch[col].str.lower().str.contains("cor",  na=False).sum()
                    break
        taxa = f"{n_prev/(n_prev+n_corr)*100:.0f}%" if (n_prev+n_corr) > 0 else "—"

        mttr_med = df_retornos["Tempo_min"].dropna().mean() if "Tempo_min" in df_retornos.columns else None

        k1.metric("PMs Cadastradas", total_pms)
        k2.metric("⛔ Atrasadas",     int(atrasadas),  delta=f"-{int(atrasadas)}" if atrasadas else None, delta_color="inverse")
        k3.metric("⚠️ Próximas 7d",   int(proximas_7))
        k4.metric("📅 Próximas 30d",   int(proximas_30))
        k5.metric("⏱ MTTR Médio",     formatar_tempo(mttr_med) if mttr_med else "—")

        st.markdown("---")

        sub1, sub2, sub3, sub4, sub5 = st.tabs(["📅 Calendário", "⚠️ Status & Alertas", "📊 Indicadores", "🤖 Mensagens Atlas", "📥 Importar do Histórico"])

        # ── Importar do histórico (disponível mesmo com preventiva já configurada) ──
        with sub5:
            _gerar_importacao_historico(df_ch, col_tipo_manut, col_maquina, col_problema,
                                        col_data_cham, col_mecanico, catalogo)

        # ── Calendário Mensal ──────────────────────────────────────────────────
        with sub1:
            hoje = datetime.today()
            ca, cb = st.columns([1, 3])
            with ca:
                ano_cal = st.selectbox("Ano",  list(range(hoje.year-1, hoje.year+3)), index=1)
                mes_cal = st.selectbox("Mês",  list(range(1,13)),
                                       index=hoje.month-1,
                                       format_func=lambda m: ["Jan","Fev","Mar","Abr","Mai","Jun",
                                                               "Jul","Ago","Set","Out","Nov","Dez"][m-1])
            with cb:
                fig_cal = calendario_mensal(df_prev, ano_cal, mes_cal)
                st.plotly_chart(fig_cal, use_container_width=True)

            # PMs do mês selecionado
            if "Próxima_Execução" in df_prev.columns:
                pm_mes = df_prev[
                    (df_prev["Próxima_Execução"].dt.year  == ano_cal) &
                    (df_prev["Próxima_Execução"].dt.month == mes_cal)
                ].copy()
                if not pm_mes.empty:
                    pm_mes["Dia"] = pm_mes["Próxima_Execução"].dt.day
                    cols_m = [c for c in ["Dia","Patrimônio","Tarefa","Tipo","Responsável","Prioridade","Status"] if c in pm_mes.columns]
                    st.dataframe(pm_mes[cols_m].sort_values("Dia"), use_container_width=True, hide_index=True)
                else:
                    st.info("Nenhuma PM programada para este mês.")

        # ── Status & Alertas ───────────────────────────────────────────────────
        with sub2:
            if "Status" in df_prev.columns:
                # Atrasadas
                atras = df_prev[df_prev["Dias_Restantes"] < 0].copy() if "Dias_Restantes" in df_prev.columns else pd.DataFrame()
                if not atras.empty:
                    st.error(f"### ⛔ {len(atras)} PM(s) Atrasada(s)")
                    cols_a = [c for c in ["Patrimônio","Tarefa","Última_Execução","Próxima_Execução","Dias_Restantes","Responsável","Prioridade"] if c in atras.columns]
                    st.dataframe(atras[cols_a].sort_values("Dias_Restantes"), use_container_width=True, hide_index=True)

                # Urgentes (dentro do prazo de antecedência)
                urg = df_prev[(df_prev["Dias_Restantes"] >= 0) & (df_prev["Dias_Restantes"] <= antecedencia)].copy() if "Dias_Restantes" in df_prev.columns else pd.DataFrame()
                if not urg.empty:
                    st.warning(f"### ⚠️ {len(urg)} PM(s) Vencendo em {antecedencia} dias")
                    cols_u = [c for c in ["Patrimônio","Tarefa","Próxima_Execução","Dias_Restantes","Responsável","Prioridade"] if c in urg.columns]
                    st.dataframe(urg[cols_u].sort_values("Dias_Restantes"), use_container_width=True, hide_index=True)

                # Todas as PMs com status
                st.markdown("---")
                st.subheader("📋 Todas as PMs")
                cols_t = [c for c in ["Status","Patrimônio","Tarefa","Tipo","Frequência_Dias","Última_Execução","Próxima_Execução","Responsável","Prioridade","Observações"] if c in df_prev.columns]
                st.dataframe(
                    df_prev[cols_t].sort_values("Dias_Restantes") if "Dias_Restantes" in df_prev.columns else df_prev[cols_t],
                    use_container_width=True, hide_index=True
                )

        # ── Indicadores PCM ───────────────────────────────────────────────────
        with sub3:
            st.subheader("📊 Indicadores de Manutenção (PCM Industrial)")

            col_i1, col_i2 = st.columns(2)

            with col_i1:
                # MTBF por máquina
                if not df_mtbf.empty and "MTBF_h" in df_mtbf.columns:
                    st.markdown("#### ⏳ MTBF – Tempo Médio Entre Falhas")
                    mtbf_df = df_mtbf[["Máquina","MTBF_h"]].dropna().sort_values("MTBF_h")
                    mtbf_df["MTBF"] = mtbf_df["MTBF_h"].apply(
                        lambda h: f"{int(h//24)}d {int(h%24)}h" if h >= 24 else f"{int(h)}h"
                    )
                    fig_mtbf = px.bar(mtbf_df.head(20), x="MTBF_h", y="Máquina",
                                      orientation="h", text="MTBF",
                                      color="MTBF_h", color_continuous_scale="RdYlGn",
                                      title="MTBF por Máquina (maior = melhor)")
                    fig_mtbf.update_layout(xaxis_title="Horas", yaxis_title="",
                                           yaxis={"categoryorder":"total ascending"})
                    st.plotly_chart(fig_mtbf, use_container_width=True)
                else:
                    st.info("MTBF disponível após acumular chamados históricos.")

            with col_i2:
                # MTTR por máquina
                if not df_mtbf.empty and "MTTR_min" in df_mtbf.columns:
                    st.markdown("#### 🔧 MTTR – Tempo Médio de Reparo")
                    mttr_df = df_mtbf[["Máquina","MTTR_min"]].dropna().sort_values("MTTR_min", ascending=False)
                    mttr_df["MTTR"] = mttr_df["MTTR_min"].apply(formatar_tempo)
                    fig_mttr = px.bar(mttr_df.head(20), x="MTTR_min", y="Máquina",
                                      orientation="h", text="MTTR",
                                      color="MTTR_min", color_continuous_scale="RdYlGn_r",
                                      title="MTTR por Máquina (menor = melhor)")
                    fig_mttr.update_layout(xaxis_title="Minutos", yaxis_title="",
                                           yaxis={"categoryorder":"total ascending"})
                    st.plotly_chart(fig_mttr, use_container_width=True)
                else:
                    st.info("MTTR disponível após acumular retornos com timestamp.")

            # Conformidade de PMs
            if not df_prev.empty and "Dias_Restantes" in df_prev.columns:
                st.markdown("---")
                st.markdown("#### 📈 Conformidade de PMs")
                em_dia   = (df_prev["Dias_Restantes"] >= 0).sum()
                atrasado = (df_prev["Dias_Restantes"] < 0).sum()
                conf_pct = em_dia / len(df_prev) * 100 if len(df_prev) > 0 else 0
                fig_conf = go.Figure(go.Indicator(
                    mode="gauge+number",
                    value=conf_pct,
                    number={"suffix": "%"},
                    title={"text": "Índice de Conformidade"},
                    gauge={
                        "axis": {"range": [0, 100]},
                        "bar":  {"color": "#2ecc71" if conf_pct >= 80 else "#f39c12" if conf_pct >= 60 else "#e74c3c"},
                        "steps": [
                            {"range": [0,  60], "color": "#2d1b1b"},
                            {"range": [60, 80], "color": "#2d2510"},
                            {"range": [80,100], "color": "#1b2d1b"},
                        ],
                        "threshold": {"line": {"color": "white","width": 2}, "value": 80}
                    }
                ))
                fig_conf.update_layout(height=280, paper_bgcolor="#0e1117", font=dict(color="#fafafa"))
                st.plotly_chart(fig_conf, use_container_width=True)

        # ── Mensagens Atlas ────────────────────────────────────────────────────
        with sub4:
            st.subheader("🤖 Gerar Mensagens para o Bot Atlas")
            st.caption(f"Mostrando PMs atrasadas + vencendo em até {antecedencia} dias")

            if "Dias_Restantes" not in df_prev.columns:
                st.info("Configure a planilha preventiva para gerar mensagens.")
            else:
                alerta_df = df_prev[df_prev["Dias_Restantes"] <= antecedencia].copy()
                if alerta_df.empty:
                    st.success("✅ Nenhuma PM requer alerta no momento!")
                else:
                    msgs = []
                    for _, row in alerta_df.iterrows():
                        msgs.append(gerar_msg_telegram(row, antecedencia))

                    msg_completa = "\n\n──────────────────\n\n".join(msgs)

                    st.text_area("📋 Copie e envie pelo Telegram", msg_completa, height=400)
                    st.download_button(
                        "⬇ Baixar mensagens (.txt)",
                        msg_completa.encode("utf-8"),
                        file_name=f"alertas_pm_{datetime.today().strftime('%Y%m%d')}.txt",
                        mime="text/plain"
                    )

                    st.markdown("---")
                    st.info("💡 **Dica**: Copie o texto acima e envie diretamente ao bot Atlas pelo Telegram, ou programe o Atlas para enviar automaticamente usando a lista de datas da planilha.")


# ─────────────────────────────────────────────
# ABA 6: Relatórios
# ─────────────────────────────────────────────
with tab_relatorio:
    st.markdown("## 📄 Gerador de Relatórios")
    st.caption("Relatórios exportados em Excel (.xlsx) com formatação profissional MassaFest.")

    tipo_rel = st.radio(
        "Selecione o relatório",
        ["📅 Relatório Mensal Geral", "🔧 PMs Preventivas do Mês"],
        horizontal=True
    )

    st.markdown("---")

    # ── Seletor de período ─────────────────────────────────────────────────
    col_pa, col_pb = st.columns(2)
    with col_pa:
        mes_rel = st.selectbox("Mês", list(range(1, 13)), index=datetime.today().month - 1,
                               format_func=lambda m: ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                                                       "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"][m-1])
    with col_pb:
        ano_rel = st.selectbox("Ano", list(range(datetime.today().year - 2, datetime.today().year + 1)),
                               index=2)

    ini_rel = pd.Timestamp(ano_rel, mes_rel, 1)
    fim_rel = (ini_rel + pd.offsets.MonthEnd(0))
    st.caption(f"Período: {ini_rel.strftime('%d/%m/%Y')} a {fim_rel.strftime('%d/%m/%Y')}")

    def gerar_excel_mensal(df_ch_f, df_ret_f, col_maq, col_mec, col_data_c, col_data_r, mes, ano, catalogo):
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference

        wb = openpyxl.Workbook()

        LARANJA = "F97316"
        ESCURO  = "1A1A1A"
        BRANCO  = "FFFFFF"
        CINZA   = "F5F5F5"
        BORDA   = Side(style="thin", color="DDDDDD")

        def estilo_header(ws, row, cols):
            for c in range(1, cols + 1):
                cel = ws.cell(row=row, column=c)
                cel.fill      = PatternFill("solid", fgColor=LARANJA)
                cel.font      = Font(bold=True, color=BRANCO, size=11)
                cel.alignment = Alignment(horizontal="center", vertical="center")
                cel.border    = Border(bottom=Side(style="medium", color=LARANJA))

        def estilo_linha(ws, row, cols, zebra=False):
            for c in range(1, cols + 1):
                cel = ws.cell(row=row, column=c)
                if zebra:
                    cel.fill = PatternFill("solid", fgColor="FFF7ED")
                cel.border = Border(
                    top=BORDA, bottom=BORDA, left=BORDA, right=BORDA
                )
                cel.alignment = Alignment(vertical="center")

        def titulo_planilha(ws, texto, subtexto=""):
            ws.merge_cells("A1:H1")
            ws["A1"] = "🏭  MASSAFEST  –  PCM"
            ws["A1"].font      = Font(bold=True, size=16, color=LARANJA)
            ws["A1"].fill      = PatternFill("solid", fgColor=ESCURO)
            ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[1].height = 36

            ws.merge_cells("A2:H2")
            ws["A2"] = texto
            ws["A2"].font      = Font(bold=True, size=13, color=BRANCO)
            ws["A2"].fill      = PatternFill("solid", fgColor="2D2D2D")
            ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[2].height = 28

            if subtexto:
                ws.merge_cells("A3:H3")
                ws["A3"] = subtexto
                ws["A3"].font      = Font(italic=True, size=10, color="AAAAAA")
                ws["A3"].fill      = PatternFill("solid", fgColor="1A1A1A")
                ws["A3"].alignment = Alignment(horizontal="left")
                ws.row_dimensions[3].height = 20

        nome_mes = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                    "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"][mes-1]

        # ── ABA 1: Resumo ────────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Resumo"
        ws1.sheet_view.showGridLines = False
        titulo_planilha(ws1, f"Relatório Mensal de Manutenção – {nome_mes}/{ano}",
                        f"Gerado em {datetime.today().strftime('%d/%m/%Y às %H:%M')}")

        kpis = [
            ("Total de Chamados",      len(df_ch_f),  ""),
            ("Atendimentos Realizados",len(df_ret_f),  ""),
            ("Máquinas Atendidas",
             df_ch_f[col_maq].nunique() if col_maq != "(nenhuma)" and not df_ch_f.empty else 0, ""),
            ("Mecânicos Ativos",
             df_ret_f[col_mec].nunique() if col_mec != "(nenhuma)" and not df_ret_f.empty else 0, ""),
            ("Tempo Médio de Atendimento",
             formatar_tempo(df_ret_f["Tempo_min"].dropna().mean()) if "Tempo_min" in df_ret_f.columns else "—", ""),
        ]
        ws1.row_dimensions[5].height = 20
        headers_kpi = ["Indicador", "Valor", ""]
        for c, h in enumerate(headers_kpi, 1):
            ws1.cell(5, c, h)
        estilo_header(ws1, 5, 3)

        for i, (ind, val, _) in enumerate(kpis, 6):
            ws1.cell(i, 1, ind).font = Font(bold=True, size=11)
            ws1.cell(i, 2, val)
            ws1.cell(i, 2).alignment = Alignment(horizontal="center")
            estilo_linha(ws1, i, 3, zebra=(i % 2 == 0))
            ws1.row_dimensions[i].height = 22

        ws1.column_dimensions["A"].width = 35
        ws1.column_dimensions["B"].width = 18

        # ── ABA 2: Chamados ──────────────────────────────────────────────
        ws2 = wb.create_sheet("Chamados")
        ws2.sheet_view.showGridLines = False
        titulo_planilha(ws2, f"Chamados – {nome_mes}/{ano}")

        if not df_ch_f.empty:
            colunas_ch = list(df_ch_f.columns)[:8]
            for c, col in enumerate(colunas_ch, 1):
                ws2.cell(5, c, str(col))
            estilo_header(ws2, 5, len(colunas_ch))

            for i, (_, row) in enumerate(df_ch_f[colunas_ch].iterrows(), 6):
                for c, val in enumerate(row.values, 1):
                    v = val.strftime("%d/%m/%Y %H:%M") if isinstance(val, pd.Timestamp) else str(val) if str(val) != "nan" else ""
                    ws2.cell(i, c, v)
                estilo_linha(ws2, i, len(colunas_ch), zebra=(i % 2 == 0))
                ws2.row_dimensions[i].height = 20

            for c in range(1, len(colunas_ch) + 1):
                ws2.column_dimensions[get_column_letter(c)].width = 22

        # ── ABA 3: Mecânicos ────────────────────────────────────────────
        ws3 = wb.create_sheet("Mecânicos")
        ws3.sheet_view.showGridLines = False
        titulo_planilha(ws3, f"Desempenho de Mecânicos – {nome_mes}/{ano}")

        if not df_ret_f.empty and col_mec != "(nenhuma)" and col_mec in df_ret_f.columns:
            mec_df = df_ret_f[col_mec].value_counts().reset_index()
            mec_df.columns = ["Mecânico", "Atendimentos"]
            mec_df["% do Total"] = (mec_df["Atendimentos"] / mec_df["Atendimentos"].sum() * 100).round(1)
            if "Tempo_min" in df_ret_f.columns:
                tm = df_ret_f.groupby(col_mec)["Tempo_min"].mean().reset_index()
                tm.columns = ["Mecânico", "Tempo_Médio_min"]
                mec_df = mec_df.merge(tm, on="Mecânico", how="left")
                mec_df["Tempo Médio"] = mec_df["Tempo_Médio_min"].apply(formatar_tempo)
                mec_df = mec_df.drop(columns=["Tempo_Médio_min"])

            cols_m3 = list(mec_df.columns)
            for c, col in enumerate(cols_m3, 1):
                ws3.cell(5, c, col)
            estilo_header(ws3, 5, len(cols_m3))

            for i, (_, row) in enumerate(mec_df.iterrows(), 6):
                for c, val in enumerate(row.values, 1):
                    ws3.cell(i, c, val if str(val) != "nan" else "")
                    ws3.cell(i, c).alignment = Alignment(horizontal="center")
                estilo_linha(ws3, i, len(cols_m3), zebra=(i % 2 == 0))
                ws3.row_dimensions[i].height = 22

            for c in range(1, len(cols_m3) + 1):
                ws3.column_dimensions[get_column_letter(c)].width = 20

        # ── ABA 4: Máquinas ─────────────────────────────────────────────
        ws4 = wb.create_sheet("Máquinas")
        ws4.sheet_view.showGridLines = False
        titulo_planilha(ws4, f"Máquinas com mais Manutenção – {nome_mes}/{ano}")

        if not df_ch_f.empty and col_maq != "(nenhuma)" and col_maq in df_ch_f.columns:
            maq_cnt = df_ch_f[col_maq].value_counts().reset_index()
            maq_cnt.columns = ["Máquina", "Chamados"]
            maq_cnt["Nome"] = maq_cnt["Máquina"].apply(
                lambda m: catalogo.get(str(m).upper(), {}).get("nome", "")
            )
            maq_cnt["Localização"] = maq_cnt["Máquina"].apply(
                lambda m: catalogo.get(str(m).upper(), {}).get("localizacao", "")
            )

            cols_m4 = ["Máquina", "Nome", "Localização", "Chamados"]
            for c, col in enumerate(cols_m4, 1):
                ws4.cell(5, c, col)
            estilo_header(ws4, 5, len(cols_m4))

            for i, (_, row) in enumerate(maq_cnt.iterrows(), 6):
                for c, col in enumerate(cols_m4, 1):
                    ws4.cell(i, c, str(row[col]) if str(row[col]) != "nan" else "")
                estilo_linha(ws4, i, len(cols_m4), zebra=(i % 2 == 0))
                ws4.row_dimensions[i].height = 20

            ws4.column_dimensions["A"].width = 12
            ws4.column_dimensions["B"].width = 40
            ws4.column_dimensions["C"].width = 25
            ws4.column_dimensions["D"].width = 12

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def gerar_excel_preventiva(df_prev_f, mes, ano, catalogo):
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb   = openpyxl.Workbook()
        ws   = wb.active
        ws.title = "PMs do Mês"
        ws.sheet_view.showGridLines = False

        LARANJA = "F97316"; ESCURO = "1A1A1A"; BRANCO = "FFFFFF"
        VERDE = "22C55E"; VERMELHO = "EF4444"; AMARELO = "EAB308"
        BORDA = Side(style="thin", color="DDDDDD")

        nome_mes = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                    "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"][mes-1]

        ws.merge_cells("A1:I1")
        ws["A1"] = "🏭  MASSAFEST  –  PCM"
        ws["A1"].font      = Font(bold=True, size=16, color=LARANJA)
        ws["A1"].fill      = PatternFill("solid", fgColor=ESCURO)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 36

        ws.merge_cells("A2:I2")
        ws["A2"] = f"Manutenções Preventivas – {nome_mes}/{ano}"
        ws["A2"].font      = Font(bold=True, size=13, color=BRANCO)
        ws["A2"].fill      = PatternFill("solid", fgColor="2D2D2D")
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 28

        headers = ["Patrimônio","Nome Máquina","Tarefa","Tipo","Frequência","Próxima PM","Responsável","Prioridade","Status"]
        for c, h in enumerate(headers, 1):
            cel = ws.cell(4, c, h)
            cel.fill      = PatternFill("solid", fgColor=LARANJA)
            cel.font      = Font(bold=True, color=BRANCO, size=11)
            cel.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[4].height = 24

        cor_status = {"Atrasada": VERMELHO, "Urgente": AMARELO, "OK": VERDE}

        row_num = 5
        if not df_prev_f.empty and "Próxima_Execução" in df_prev_f.columns:
            pm_mes = df_prev_f[
                (df_prev_f["Próxima_Execução"].dt.year  == ano) &
                (df_prev_f["Próxima_Execução"].dt.month == mes)
            ].copy() if not df_prev_f.empty else pd.DataFrame()

            df_usar = pm_mes if not pm_mes.empty else df_prev_f

            for _, row in df_usar.iterrows():
                pat  = str(row.get("Patrimônio",""))
                nome = catalogo.get(pat.upper(), {}).get("nome", "")
                data = row["Próxima_Execução"].strftime("%d/%m/%Y") if pd.notna(row.get("Próxima_Execução")) else ""
                dias = row.get("Dias_Restantes", None)
                status = "Atrasada" if (dias is not None and dias < 0) else ("Urgente" if (dias is not None and dias <= 7) else "OK")

                vals = [pat, nome, row.get("Tarefa",""), row.get("Tipo",""),
                        f"{int(row.get('Frequência_Dias',0))} dias" if row.get("Frequência_Dias") else "",
                        data, row.get("Responsável",""), row.get("Prioridade",""), status]

                for c, v in enumerate(vals, 1):
                    cel = ws.cell(row_num, c, v)
                    cel.alignment = Alignment(vertical="center")
                    cel.border    = Border(top=BORDA, bottom=BORDA, left=BORDA, right=BORDA)
                    if c == 9:  # coluna Status
                        cor = cor_status.get(status, "888888")
                        cel.fill = PatternFill("solid", fgColor=cor)
                        cel.font = Font(bold=True, color=BRANCO)
                        cel.alignment = Alignment(horizontal="center", vertical="center")
                    elif row_num % 2 == 0:
                        cel.fill = PatternFill("solid", fgColor="FFF7ED")

                ws.row_dimensions[row_num].height = 22
                row_num += 1

        larguras = [12, 38, 28, 16, 12, 14, 16, 12, 12]
        for c, w in enumerate(larguras, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ── Preview e botão de download ────────────────────────────────────────
    if tipo_rel == "📅 Relatório Mensal Geral":
        df_ch_rel  = df_ch[(df_ch[col_data_cham]  >= ini_rel) & (df_ch[col_data_cham]  <= fim_rel)] if col_data_cham  != "(nenhuma)" and not df_ch.empty  else df_ch
        df_ret_rel = df_ret[(df_ret[col_data_ret] >= ini_rel) & (df_ret[col_data_ret] <= fim_rel)] if col_data_ret != "(nenhuma)" and not df_ret.empty else df_ret

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Chamados no período",   len(df_ch_rel))
        c2.metric("Atendimentos",          len(df_ret_rel))
        c3.metric("Máquinas",              df_ch_rel[col_maquina].nunique() if col_maquina != "(nenhuma)" and not df_ch_rel.empty else 0)
        c4.metric("Mecânicos",             df_ret_rel[col_mecanico].nunique() if col_mecanico != "(nenhuma)" and not df_ret_rel.empty else 0)

        st.markdown("---")
        if st.button("⬇ Gerar Relatório Mensal Excel", type="primary", use_container_width=True):
            with st.spinner("Gerando relatório…"):
                buf = gerar_excel_mensal(df_ch_rel, df_ret_rel, col_maquina, col_mecanico,
                                         col_data_cham, col_data_ret, mes_rel, ano_rel, catalogo)
            nome_mes_str = ["jan","fev","mar","abr","mai","jun","jul","ago","set","out","nov","dez"][mes_rel-1]
            st.download_button(
                label="📥 Baixar Excel",
                data=buf,
                file_name=f"PCM_MassaFest_{nome_mes_str}{ano_rel}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    else:  # PMs Preventivas
        if df_prev.empty:
            st.info("Configure a planilha de PMs preventivas no menu lateral para gerar este relatório.")
        else:
            pm_mes_cnt = len(df_prev[
                (df_prev["Próxima_Execução"].dt.year  == ano_rel) &
                (df_prev["Próxima_Execução"].dt.month == mes_rel)
            ]) if "Próxima_Execução" in df_prev.columns else 0

            st.metric("PMs programadas no período", pm_mes_cnt)
            st.markdown("---")

            if st.button("⬇ Gerar Relatório de PMs Excel", type="primary", use_container_width=True):
                with st.spinner("Gerando relatório…"):
                    buf = gerar_excel_preventiva(df_prev, mes_rel, ano_rel, catalogo)
                nome_mes_str = ["jan","fev","mar","abr","mai","jun","jul","ago","set","out","nov","dez"][mes_rel-1]
                st.download_button(
                    label="📥 Baixar Excel",
                    data=buf,
                    file_name=f"PCM_MassaFest_Preventiva_{nome_mes_str}{ano_rel}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )


# ABA 5: Dados Brutos
# ─────────────────────────────────────────────
with tab_dados:
    st.subheader("Dados brutos")
    aba = st.radio("Planilha", ["Chamados", "Retornos"], horizontal=True)
    df_show = df_ch if aba == "Chamados" else df_ret
    if df_show.empty:
        st.info("Nenhum dado disponível.")
    else:
        st.write(f"{len(df_show)} registros")
        st.dataframe(df_show, use_container_width=True)
        csv = df_show.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            label="⬇ Baixar CSV",
            data=csv,
            file_name=f"{aba.lower()}_{datetime.today().strftime('%Y%m%d')}.csv",
            mime="text/csv"
        )
